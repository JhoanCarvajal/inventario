from django.http import response
from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from openpyxl import Workbook
from django.http import request
from django.http.response import HttpResponse
from products.models import Product

# Create your views here.
# landing page
class LandingPageView(TemplateView):
    template_name = "core/landing_page.html"


# generar excel con datos
class ReporteExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        products = Product.objects.filter(id_usuario=self.request.user)
        wb = Workbook()
        ws = wb.active

        ws['B1'] = "Productos"
        ws.merge_cells('B1:C1')
        
        ws['A2'] = "Nombre"
        ws['B2'] = "Precio"
        ws['C2'] = "Información"
        ws['D2'] = "Fecha"

        cont = 3

        for product in products:
            ws.cell(row = cont, column = 1).value = product.nombre
            ws.cell(row = cont, column = 2).value = product.precio
            ws.cell(row = cont, column = 3).value = product.informacion
            ws.cell(row = cont, column = 4).value = str(product.fecha_adquisicion)
            cont += 1
        
        nombre_archivo = "productos.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)

        return response